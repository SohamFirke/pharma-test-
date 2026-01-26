"""
Excel Generator Utility for Admin Dashboard Exports

Generates professionally formatted Excel files for:
1. Product List (Medicine Catalog)
2. Consumer Order History

For administrative and inventory management purposes only.
No medical decision-making logic involved.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from pathlib import Path
from datetime import datetime
from typing import Dict, Any
import random
import sys

# Add parent directory to path
sys.path.append(str(Path(__file__).resolve().parent.parent))

from database import Database


class ExcelGenerator:
    """
    Generate professionally formatted Excel files for admin exports.
    
    Purpose: Administrative and inventory management only.
    Authority: No medical diagnosis or decision-making logic.
    """
    
    # Mock data generators for demonstration
    GENDERS = ["Male", "Female", "Other"]
    DOSAGE_FREQUENCIES = [
        "Once daily",
        "Twice daily", 
        "Three times daily",
        "As needed",
        "Every 4-6 hours",
        "Before meals"
    ]
    
    # Medicine categories
    CATEGORIES = {
        "pain": "Pain Relief",
        "antibiotic": "Antibiotic",
        "diabetes": "Diabetes Management",
        "blood pressure": "Cardiovascular",
        "allergy": "Allergy Relief",
        "default": "General Medicine"
    }
    
    # Mock descriptions
    DESCRIPTIONS = {
        "Paracetamol": "Pain relief and fever reducer",
        "Ibuprofen": "Anti-inflammatory and pain relief",
        "Amoxicillin": "Antibiotic for bacterial infections",
        "Lisinopril": "Blood pressure management",
        "Metformin": "Type 2 diabetes management",
        "Insulin Glargine": "Long-acting insulin for diabetes",
        "Aspirin": "Pain relief and blood thinner",
        "Omeprazole": "Acid reflux and heartburn treatment",
        "Simvastatin": "Cholesterol management",
        "Levothyroxine": "Thyroid hormone replacement"
    }
    
    @staticmethod
    def _apply_header_formatting(worksheet, num_columns: int):
        """Apply professional formatting to header row."""
        # Bold font for headers
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    @staticmethod
    def _auto_size_columns(worksheet):
        """Auto-size all columns for readability."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _get_category(medicine_name: str) -> str:
        """Determine category based on medicine name."""
        name_lower = medicine_name.lower()
        
        for keyword, category in ExcelGenerator.CATEGORIES.items():
            if keyword in name_lower:
                return category
        
        return ExcelGenerator.CATEGORIES["default"]
    
    @staticmethod
    def _get_description(medicine_name: str) -> str:
        """Get description for medicine."""
        return ExcelGenerator.DESCRIPTIONS.get(
            medicine_name, 
            "Pharmaceutical product for medical treatment"
        )
    
    @staticmethod
    def _generate_price(medicine_name: str, index: int) -> float:
        """Generate realistic price based on medicine type."""
        # Prescription meds typically more expensive
        if any(keyword in medicine_name.lower() for keyword in ["insulin", "antibiotic", "prescription"]):
            base = 50.0
            variance = 100.0
        else:
            base = 10.0
            variance = 40.0
        
        # Use index as seed for consistency
        random.seed(hash(medicine_name) % 10000)
        return round(base + random.random() * variance, 2)
    
    @classmethod
    def generate_product_list(cls) -> bytes:
        """
        Generate Product List Excel file.
        
        Columns:
        1. Product ID
        2. Product Name
        3. PZN/SKU
        4. Price
        5. Package Size
        6. Category
        7. Description
        8. Prescription Required
        9. Stock Quantity
        10. Status
        
        Returns:
            Excel file as bytes (in-memory)
        """
        try:
            # Load medicine data
            medicines_df = Database.load_medicine_master()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Define headers (EXACT ORDER as specified)
            headers = [
                "Product ID",
                "Product Name",
                "PZN / SKU",
                "Price",
                "Package Size",
                "Category",
                "Description",
                "Prescription Required",
                "Stock Quantity",
                "Status"
            ]
            
            ws.append(headers)
            
            # Populate data
            for idx, row in medicines_df.iterrows():
                product_id = idx + 1
                medicine_name = row['medicine_name']
                pzn_sku = f"MED-{product_id:04d}"
                price = cls._generate_price(medicine_name, idx)
                package_size = f"{row['unit_type']}s"
                category = cls._get_category(medicine_name)
                description = cls._get_description(medicine_name)
                prescription_req = "Yes" if row.get('prescription_required', False) else "No"
                stock_qty = int(row['stock_level'])
                status = "Active" if stock_qty > 0 else "Out of Stock"
                
                ws.append([
                    product_id,
                    medicine_name,
                    pzn_sku,
                    price,
                    package_size,
                    category,
                    description,
                    prescription_req,
                    stock_qty,
                    status
                ])
            
            # Apply formatting
            cls._apply_header_formatting(ws, len(headers))
            
            # Format Price column (column D, index 4)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=4).number_format = '$#,##0.00'
            
            # Auto-size columns
            cls._auto_size_columns(ws)
            
            # Save to bytes
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
        
        except Exception as e:
            print(f"Error generating product list: {e}")
            raise
    
    @classmethod
    def generate_order_history(cls) -> bytes:
        """
        Generate Consumer Order History Excel file.
        
        Columns:
        1. Order ID
        2. Patient ID
        3. Patient Age
        4. Patient Gender
        5. Order Date
        6. Product Name
        7. Quantity
        8. Dosage Frequency
        9. Prescription Required
        10. Total Price
        11. Order Status
        
        Returns:
            Excel file as bytes (in-memory)
        """
        try:
            # Load order data using Database method
            orders_df = Database.load_order_history()
            
            if orders_df.empty:
                # Create empty workbook with headers only
                wb = Workbook()
                ws = wb.active
                ws.title = "Orders"
                
                headers = [
                    "Order ID",
                    "Patient ID",
                    "Patient Age",
                    "Patient Gender",
                    "Order Date",
                    "Product Name",
                    "Quantity",
                    "Dosage Frequency",
                    "Prescription Required",
                    "Total Price",
                    "Order Status"
                ]
                
                ws.append(headers)
                cls._apply_header_formatting(ws, len(headers))
                
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer.getvalue()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders"
            
            # Define headers (EXACT ORDER as specified)
            headers = [
                "Order ID",
                "Patient ID",
                "Patient Age",
                "Patient Gender",
                "Order Date",
                "Product Name",
                "Quantity",
                "Dosage Frequency",
                "Prescription Required",
                "Total Price",
                "Order Status"
            ]
            
            ws.append(headers)
            
            # Populate data
            for idx, row in orders_df.iterrows():
                # Generate order ID
                order_id = f"ORD-{idx+1:04d}"
                
                patient_id = row['user_id']
                
                # Generate mock demographic data (for demonstration)
                random.seed(hash(patient_id) % 10000)
                patient_age = random.randint(18, 80)
                patient_gender = random.choice(cls.GENDERS)
                
                # Format order date
                try:
                    order_date = pd.to_datetime(row['purchase_date']).strftime('%Y-%m-%d')
                except:
                    order_date = datetime.now().strftime('%Y-%m-%d')
                
                product_name = row['medicine_name']
                quantity = int(row['quantity'])
                
                # Generate dosage frequency (mock)
                random.seed(hash(order_id) % 10000)
                dosage_freq = random.choice(cls.DOSAGE_FREQUENCIES)
                
                # Get prescription requirement
                medicine_info = Database.get_medicine(product_name)
                prescription_req = "Yes" if medicine_info and medicine_info.get('prescription_required', False) else "No"
                
                # Calculate total price
                unit_price = cls._generate_price(product_name, idx)
                total_price = unit_price * quantity
                
                # Order status (default)
                order_status = "Completed"
                
                ws.append([
                    order_id,
                    patient_id,
                    patient_age,
                    patient_gender,
                    order_date,
                    product_name,
                    quantity,
                    dosage_freq,
                    prescription_req,
                    total_price,
                    order_status
                ])
            
            # Apply formatting
            cls._apply_header_formatting(ws, len(headers))
            
            # Format Total Price column (column J, index 10)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=10).number_format = '$#,##0.00'
            
            # Auto-size columns
            cls._auto_size_columns(ws)
            
            # Save to bytes
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
        
        except Exception as e:
            print(f"Error generating order history: {e}")
            raise
